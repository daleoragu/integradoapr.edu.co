# notas/views/import_views.py
import csv
import io
from datetime import datetime
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.contrib.auth.models import User, Group
from django.db import transaction, IntegrityError
from django.http import HttpResponseNotFound
from django.utils.text import slugify
from unidecode import unidecode
# Se importa make_password para hashear contraseñas manualmente antes de la creación en bloque
from django.contrib.auth.hashers import make_password

try:
    from openpyxl import load_workbook
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False

from ..models.perfiles import Estudiante, Docente, Curso, FichaEstudiante
from ..models.academicos import Materia, AreaConocimiento, PonderacionAreaMateria

def es_superusuario(user):
    return user.is_superuser

@login_required
@user_passes_test(es_superusuario)
def importacion_vista(request):
    """
    Gestiona la carga y el procesamiento de archivos para la carga masiva de datos,
    asegurando que todos los datos estén asociados con el colegio actual.
    """
    if not request.colegio:
        return HttpResponseNotFound("<h1>Colegio no configurado</h1>")

    if request.method == 'POST' and 'archivo_importacion' in request.FILES:
        tipo_importacion = request.POST.get('tipo_importacion')
        archivo = request.FILES['archivo_importacion']

        try:
            # Usamos transaction.atomic() para asegurar que toda la operación
            # sea un éxito o un fracaso, manteniendo la integridad de los datos.
            with transaction.atomic():
                if tipo_importacion == 'estudiantes':
                    if not EXCEL_SUPPORT:
                        raise Exception("La librería 'openpyxl' es necesaria. Instálela con 'pip install openpyxl'.")
                    if not archivo.name.endswith('.xlsx'):
                        raise Exception("Para importar estudiantes, seleccione un archivo Excel válido (.xlsx).")
                    _procesar_excel_estudiantes_optimizado(request, archivo, request.colegio)
                
                elif tipo_importacion == 'materias':
                    if not EXCEL_SUPPORT:
                        raise Exception("La librería 'openpyxl' es necesaria. Instálela con 'pip install openpyxl'.")
                    if not archivo.name.endswith('.xlsx'):
                         raise Exception(f"Para importar materias, seleccione un archivo Excel válido (.xlsx).")
                    _procesar_excel_materias(request, archivo, request.colegio)

                elif tipo_importacion == 'docentes':
                    messages.warning(request, "La importación de docentes aún no está implementada para Excel.")
                    pass
                else:
                    raise Exception("El tipo de importación seleccionado no es válido.")

        except IntegrityError as e:
            messages.error(request, f"Error de integridad: Un dato (como un documento o username) podría ya existir. Detalles: {e}")
        except Exception as e:
            messages.error(request, f"Error durante el proceso: {e}")

        return redirect(request.META.get('HTTP_REFERER', 'notas:admin_dashboard'))

    return redirect('notas:admin_dashboard')

def _procesar_excel_estudiantes_optimizado(request, archivo, colegio):
    """
    Lógica OPTIMIZADA para procesar el archivo Excel de estudiantes de forma segura y eficiente.
    Utiliza bulk_create para minimizar las consultas a la base de datos y evitar timeouts.
    """
    creados, errores, omitidos = 0, 0, 0
    grupo_estudiantes, _ = Group.objects.get_or_create(name="Estudiantes")
    
    wb = load_workbook(archivo, data_only=True)
    sheet = wb.active
    map_tipo_doc = {v.upper(): k for k, v in FichaEstudiante.TIPO_DOCUMENTO_CHOICES}
    map_grupo_sang = {v: k for k, v in FichaEstudiante.GRUPO_SANGUINEO_CHOICES}
    
    # --- PASO 1: Leer todos los datos y preparar los objetos User en memoria ---
    usuarios_a_crear = []
    datos_filas = [] # Guardaremos los datos de cada fila para el segundo paso
    
    # Obtenemos todos los usernames existentes para evitar colisiones en una sola consulta
    existing_usernames = set(User.objects.values_list('username', flat=True))

    for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
        if not any(row) or not row[0] or not row[1]:
            continue

        row_data = list(row) + [None] * (20 - len(row))
        nombres, apellidos, *_ = row_data[:20]

        try:
            # Generación de un nombre de usuario único
            primer_nombre = unidecode(str(nombres).split(' ')[0].lower())
            primer_apellido = unidecode(str(apellidos).split(' ')[0].lower())
            username_base = f"{slugify(primer_nombre)}.{slugify(primer_apellido)}"
            username_final = username_base
            counter = 1
            while username_final in existing_usernames:
                username_final = f"{username_base}{counter}"
                counter += 1
            
            # Añadimos el nuevo username al set para que no se repita en este mismo lote
            existing_usernames.add(username_final)

            # Preparamos el objeto User.
            # IMPORTANTE: bulk_create no llama a .save(), por lo que debemos hashear la contraseña manualmente.
            user = User(
                username=username_final, 
                password=make_password(username_final), # Hasheamos la contraseña
                first_name=str(nombres or '').strip().upper(), 
                last_name=str(apellidos or '').strip().upper()
            )
            usuarios_a_crear.append(user)
            datos_filas.append({'username': username_final, 'data': row_data, 'fila_num': i})

        except Exception as e:
            messages.warning(request, f"Error preparando la fila {i} del Excel: {e}")
            errores += 1

    # --- PASO 2: Crear todos los usuarios en una sola consulta a la base de datos ---
    if usuarios_a_crear:
        User.objects.bulk_create(usuarios_a_crear)
        creados = len(usuarios_a_crear)

        # --- PASO 3: Recuperar los usuarios creados y crear los objetos relacionados ---
        # Creamos un mapa de username -> user_object para un acceso rápido
        usernames_creados = [d['username'] for d in datos_filas]
        mapa_usuarios = {u.username: u for u in User.objects.filter(username__in=usernames_creados)}
        
        # Asignar todos los nuevos usuarios al grupo "Estudiantes" de forma masiva
        grupo_estudiantes.user_set.add(*mapa_usuarios.values())

        for datos in datos_filas:
            fila_num = datos['fila_num']
            try:
                user_obj = mapa_usuarios.get(datos['username'])
                if not user_obj: continue # Si por alguna razón no se encontró, lo saltamos

                (nombres, apellidos, tipo_doc_str, num_doc, nombre_curso, fecha_nac_str, 
                 lugar_nac, eps, grupo_sang_str, enfermedades, nombre_padre, cel_padre,
                 nombre_madre, cel_madre, nombre_acud, cel_acud, email_acud, 
                 espera_porteria_str, colegio_ant, grado_ant) = datos['data'][:20]

                # Buscamos el curso dentro del colegio actual
                curso = Curso.objects.filter(nombre=str(nombre_curso).strip().upper(), colegio=colegio).first()
                if not curso:
                    raise ValueError(f"El curso '{nombre_curso}' no existe en este colegio.")

                # Creamos el Estudiante y la FichaEstudiante
                estudiante_obj = Estudiante.objects.create(user=user_obj, curso=curso, colegio=colegio)
                
                fecha_nacimiento = None
                if isinstance(fecha_nac_str, datetime):
                    fecha_nacimiento = fecha_nac_str.date()
                elif isinstance(fecha_nac_str, str):
                    try:
                        fecha_nacimiento = datetime.strptime(fecha_nac_str, '%Y-%m-%d').date()
                    except (ValueError, TypeError):
                        fecha_nacimiento = None

                FichaEstudiante.objects.create(
                    estudiante=estudiante_obj,
                    tipo_documento=map_tipo_doc.get(str(tipo_doc_str).strip().upper(), 'OT') if tipo_doc_str else 'OT',
                    numero_documento=str(num_doc).strip() if num_doc else None,
                    fecha_nacimiento=fecha_nacimiento,
                    lugar_nacimiento=lugar_nac,
                    eps=eps,
                    grupo_sanguineo=map_grupo_sang.get(str(grupo_sang_str).strip(), None) if grupo_sang_str else None,
                    enfermedades_alergias=enfermedades,
                    nombre_padre=nombre_padre,
                    celular_padre=cel_padre,
                    nombre_madre=nombre_madre,
                    celular_madre=cel_madre,
                    nombre_acudiente=nombre_acud,
                    celular_acudiente=cel_acud,
                    email_acudiente=email_acud,
                    espera_en_porteria=True if espera_porteria_str and 'SI' in str(espera_porteria_str).upper() else False,
                    colegio_anterior=colegio_ant,
                    grado_anterior=grado_ant
                )
            except Exception as e:
                messages.warning(request, f"Error procesando datos relacionados en la fila {fila_num}: {e}")
                errores += 1
                creados -= 1 # Restamos del contador de éxito

    messages.success(request, f"Proceso de estudiantes completado. Creados: {creados}. Errores: {errores}. Omitidos (ya existían): {omitidos}.")


def _procesar_excel_materias(request, archivo, colegio):
    """Lógica para procesar el archivo Excel de materias para el colegio actual."""
    creadas, errores, asociaciones_creadas = 0, 0, 0
    wb = load_workbook(archivo, data_only=True)
    sheet = wb.active
    for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
        try:
            if not any(row): continue
            nombre_materia, abreviatura, nombre_area = row[:3]
            
            nombre_materia = str(nombre_materia).strip().upper()
            nombre_area = str(nombre_area).strip().upper()

            if not nombre_materia or not nombre_area:
                continue

            materia_obj, materia_fue_creada = Materia.objects.get_or_create(
                nombre=nombre_materia,
                colegio=colegio,
                defaults={'abreviatura': str(abreviatura).strip().upper() if abreviatura else None}
            )
            if materia_fue_creada:
                creadas += 1

            area_obj, _ = AreaConocimiento.objects.get_or_create(nombre=nombre_area, colegio=colegio)

            _, ponderacion_fue_creada = PonderacionAreaMateria.objects.get_or_create(
                area=area_obj,
                materia=materia_obj,
                colegio=colegio,
                defaults={'peso_porcentual': 0.00}
            )
            if ponderacion_fue_creada:
                asociaciones_creadas += 1

        except Exception as e:
            messages.warning(request, f"Error en la fila {i} del Excel de materias: {e}")
            errores += 1
            
    messages.success(request, f"Proceso de materias completado. Materias creadas: {creadas}. Asociaciones a áreas creadas: {asociaciones_creadas}. Errores: {errores}.")

# La función para procesar docentes se mantiene igual si se usa en el futuro.
def _procesar_csv_docentes(request, reader, colegio):
    """Lógica para procesar el CSV de docentes para el colegio actual (preservada)."""
    creados, errores = 0, 0
    grupo_docentes, _ = Group.objects.get_or_create(name="Docentes")
    for i, row in enumerate(reader, 2):
        try:
            if not any(row): continue
            nombres, primer_apellido, segundo_apellido, documento, email, *_ = row
            if not documento: raise ValueError("El número de documento es obligatorio.")
            if User.objects.filter(username=documento).exists(): continue
            user = User.objects.create_user(username=documento, password=documento, email=email, first_name=str(nombres or '').upper(), last_name=f"{str(primer_apellido or '').upper()} {str(segundo_apellido or '').upper()}".strip())
            user.groups.add(grupo_docentes)
            Docente.objects.create(user=user, colegio=colegio)
            creados += 1
        except Exception as e:
            messages.warning(request, f"Error en la fila {i} del CSV de docentes: {e}")
            errores += 1
    messages.success(request, f"Proceso de docentes completado. Creados: {creados}. Errores: {errores}.")
