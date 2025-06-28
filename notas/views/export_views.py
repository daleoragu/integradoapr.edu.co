# notas/views/export_views.py
import csv
from django.http import HttpResponse
from django.db.models import Q
from django.contrib.auth.decorators import login_required, user_passes_test
from django.shortcuts import render, get_object_or_404

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False

from ..models import Curso, Estudiante, FichaEstudiante, Materia, AreaConocimiento

def es_superusuario(user):
    """
    Decorator check to ensure the user is a superuser.
    """
    return user.is_superuser

# ==============================================================================
# VISTAS DE EXPORTACIÓN Y PLANTILLA DE ESTUDIANTES
# ==============================================================================

@login_required
@user_passes_test(es_superusuario)
def descargar_plantilla_estudiantes(request):
    """
    Genera y ofrece para descarga una plantilla de Excel (.xlsx) completa para 
    la importación masiva de estudiantes.
    """
    if not EXCEL_SUPPORT:
        return HttpResponse("La librería 'openpyxl' es necesaria para esta función. Por favor, instálela con 'pip install openpyxl'.", status=500)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="plantilla_migracion_estudiantes.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Estudiantes"

    headers = [
        'NOMBRES', 'APELLIDOS', 'TIPO_DOCUMENTO (CC, TI, RC, CE, OT)', 'NUMERO_DOCUMENTO', 
        'NOMBRE_CURSO', 'FECHA_NACIMIENTO (YYYY-MM-DD)', 'LUGAR_NACIMIENTO', 'EPS', 
        'GRUPO_SANGUINEO (O+, O-, A+, A-, B+, B-, AB+, AB-)', 'ENFERMEDADES_ALERGIAS',
        'NOMBRE_PADRE', 'CELULAR_PADRE', 'NOMBRE_MADRE', 'CELULAR_MADRE',
        'NOMBRE_ACUDIENTE', 'CELULAR_ACUDIENTE', 'EMAIL_ACUDIENTE', 'ESPERA_EN_PORTERIA (SI/NO)',
        'COLEGIO_ANTERIOR', 'GRADO_ANTERIOR'
    ]
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
        ws.column_dimensions[get_column_letter(col_num)].width = 22

    ws['A2'] = 'EJEMPLO:'
    ws['B2'] = 'ESTUDIANTE DE PRUEBA'
    ws['C2'] = 'TI'
    ws['D2'] = '123456789'
    ws['E2'] = 'SEXTO A'
    
    wb.save(response)
    return response


@login_required
@user_passes_test(es_superusuario)
def exportar_estudiantes_excel(request):
    """
    Exporta la lista de estudiantes (filtrada o completa) a un archivo Excel
    que coincide con el formato de la plantilla de importación.
    """
    if not EXCEL_SUPPORT:
        return HttpResponse("La librería 'openpyxl' es necesaria para esta función.", status=500)

    curso_id = request.GET.get('curso', '')
    query = request.GET.get('q', '')
    
    fichas_qs = FichaEstudiante.objects.select_related('estudiante__user', 'estudiante__curso').all().order_by('estudiante__user__last_name', 'estudiante__user__first_name')

    if curso_id:
        fichas_qs = fichas_qs.filter(estudiante__curso_id=curso_id)
    if query:
        fichas_qs = fichas_qs.filter(
            Q(estudiante__user__first_name__icontains=query) | Q(estudiante__user__last_name__icontains=query)
        )

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="exportacion_estudiantes.xlsx"'
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Estudiantes Exportados"

    headers = [
        'NOMBRES', 'APELLIDOS', 'TIPO_DOCUMENTO', 'NUMERO_DOCUMENTO', 'NOMBRE_CURSO', 
        'FECHA_NACIMIENTO', 'LUGAR_NACIMIENTO', 'EPS', 'GRUPO_SANGUINEO', 'ENFERMEDADES_ALERGIAS',
        'NOMBRE_PADRE', 'CELULAR_PADRE', 'NOMBRE_MADRE', 'CELULAR_MADRE',
        'NOMBRE_ACUDIENTE', 'CELULAR_ACUDIENTE', 'EMAIL_ACUDIENTE', 'ESPERA_EN_PORTERIA',
        'COLEGIO_ANTERIOR', 'GRADO_ANTERIOR'
    ]
    
    header_font = Font(bold=True)
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header_title)
        cell.font = header_font
        ws.column_dimensions[get_column_letter(col_num)].width = 22

    for row_num, ficha in enumerate(fichas_qs, 2):
        ws.cell(row=row_num, column=1, value=ficha.estudiante.user.first_name)
        ws.cell(row=row_num, column=2, value=ficha.estudiante.user.last_name)
        ws.cell(row=row_num, column=3, value=ficha.get_tipo_documento_display())
        ws.cell(row=row_num, column=4, value=ficha.numero_documento)
        ws.cell(row=row_num, column=5, value=ficha.estudiante.curso.nombre if ficha.estudiante.curso else '')
        ws.cell(row=row_num, column=6, value=ficha.fecha_nacimiento)
        ws.cell(row=row_num, column=7, value=ficha.lugar_nacimiento)
        ws.cell(row=row_num, column=8, value=ficha.eps)
        ws.cell(row=row_num, column=9, value=ficha.grupo_sanguineo)
        ws.cell(row=row_num, column=10, value=ficha.enfermedades_alergias)
        ws.cell(row=row_num, column=11, value=ficha.nombre_padre)
        ws.cell(row=row_num, column=12, value=ficha.celular_padre)
        ws.cell(row=row_num, column=13, value=ficha.nombre_madre)
        ws.cell(row=row_num, column=14, value=ficha.celular_madre)
        ws.cell(row=row_num, column=15, value=ficha.nombre_acudiente)
        ws.cell(row=row_num, column=16, value=ficha.celular_acudiente)
        ws.cell(row=row_num, column=17, value=ficha.email_acudiente)
        ws.cell(row=row_num, column=18, value="SI" if ficha.espera_en_porteria else "NO")
        ws.cell(row=row_num, column=19, value=ficha.colegio_anterior)
        ws.cell(row=row_num, column=20, value=ficha.grado_anterior)

    wb.save(response)
    return response

# ===============================================================
# VISTAS PARA GESTIÓN DE MATERIAS
# ===============================================================

@login_required
@user_passes_test(es_superusuario)
def descargar_plantilla_materias(request):
    """
    Genera y ofrece para descarga una plantilla de Excel para la importación de materias.
    """
    if not EXCEL_SUPPORT:
        return HttpResponse("La librería 'openpyxl' es necesaria.", status=500)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="plantilla_importacion_materias.xlsx"'
    wb = Workbook()
    ws = wb.active
    ws.title = "Materias"
    headers = ['NOMBRE_MATERIA', 'ABREVIATURA', 'INTENSIDAD_HORARIA', 'NOMBRE_AREA']
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
        ws.column_dimensions[get_column_letter(col_num)].width = 30
    
    ws['A2'] = 'EJEMPLO: MATEMÁTICAS'
    ws['B2'] = 'MAT'
    ws['C2'] = '4'
    ws['D2'] = 'CIENCIAS EXACTAS'
    
    wb.save(response)
    return response

@login_required
@user_passes_test(es_superusuario)
def exportar_materias_excel(request):
    """
    Exporta la lista actual de materias y sus áreas a un archivo Excel.
    """
    if not EXCEL_SUPPORT:
        return HttpResponse("La librería 'openpyxl' es necesaria.", status=500)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="exportacion_materias.xlsx"'
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Materias Exportadas"
    
    headers = ['NOMBRE_MATERIA', 'ABREVIATURA', 'INTENSIDAD_HORARIA', 'NOMBRE_AREA']
    header_font = Font(bold=True)
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header_title)
        cell.font = header_font
        ws.column_dimensions[get_column_letter(col_num)].width = 35
    
    materias = Materia.objects.select_related('area').all().order_by('area__nombre', 'nombre')
    for row_num, materia in enumerate(materias, 2):
        ws.cell(row=row_num, column=1, value=materia.nombre)
        ws.cell(row=row_num, column=2, value=materia.abreviatura)
        ws.cell(row=row_num, column=3, value=materia.intensidad_horaria)
        ws.cell(row=row_num, column=4, value=materia.area.nombre)
        
    wb.save(response)
    return response

# ==============================================================================
# VISTAS CONSERVADAS (PARA DOCENTES Y SELECTOR GENERAL)
# ==============================================================================

@login_required
@user_passes_test(es_superusuario)
def descargar_plantilla_docentes(request):
    """
    Genera y ofrece para descarga una plantilla de Excel para la importación de docentes.
    """
    if not EXCEL_SUPPORT:
        return HttpResponse("La librería 'openpyxl' es necesaria para esta función.", status=500)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="plantilla_importacion_docentes.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Docentes"

    headers = ['NOMBRES', 'PRIMER_APELLIDO', 'SEGUNDO_APELLIDO', 'DOCUMENTO', 'EMAIL']
    header_font = Font(bold=True)
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header_title)
        cell.font = header_font
        ws.column_dimensions[get_column_letter(col_num)].width = 25

    wb.save(response)
    return response

@login_required
@user_passes_test(es_superusuario)
def selector_exportacion_vista(request):
    """
    Muestra una página con opciones para que el administrador
    pueda seleccionar qué datos exportar. (Esta vista puede ser obsoleta pero se mantiene por compatibilidad)
    """
    cursos = Curso.objects.all().order_by('nombre')
    context = {
        'cursos': cursos,
    }
    return render(request, 'notas/admin_tools/selector_exportacion.html', context)
